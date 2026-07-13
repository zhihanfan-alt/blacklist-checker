"""
分拣员工黑名单检查器 — FastAPI 后端
6层名称匹配 + 多文件批量检查 + 多分表导出
"""

import io
import re
import zipfile
from pathlib import Path
from typing import Optional
from urllib.parse import quote

import openpyxl
import uvicorn
from fastapi import FastAPI, File, Form, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from rapidfuzz import fuzz
from unidecode import unidecode

# ── 常量 ──────────────────────────────────────────────

NAME_COLUMNS = [1, 4, 8, 11, 15, 19, 23, 28]  # A D H K O S W AB
SECTIONS = {
    1: "Escaneadoras", 4: "Mesas", 8: "姓名列表",
    11: "Name/Time1", 15: "Name/Time2", 19: "Lista manana",
    23: "名字/Nombre", 28: "名字/Nombre2",
}
FUZZY_THRESHOLD = 85
NGRAM_THRESHOLD = 0.70

# ── 黑名单内存状态 ────────────────────────────────────

blacklist_names: list[str] = []  # 保留原始大小写


def _norm(name: str) -> str:
    """归一化：去重音+小写+去标点+去空格"""
    name = name.strip().lower()
    name = unidecode(name)
    name = re.sub(r"[^\w\s]", "", name)
    return re.sub(r"\s+", " ", name).strip()


def _blacklist_set() -> set[str]:
    return {_norm(n) for n in blacklist_names}


# ── 名称匹配 ──────────────────────────────────────────

def ngram_sim(a: str, b: str) -> float:
    def ng(s, n): return {s[i:i+n] for i in range(len(s) - n + 1)}
    ba, bb = ng(a, 2), ng(b, 2)
    ta, tb = ng(a, 3), ng(b, 3)
    bs = len(ba & bb) / len(ba | bb) if (ba | bb) else 0
    ts = len(ta & tb) / len(ta | tb) if (ta | tb) else 0
    return (bs + ts) / 2


def match_score(sched: str, bl: str) -> tuple[int, str, str]:
    if not sched or not bl:
        return 0, "", ""
    a, b = _norm(sched), _norm(bl)
    if not a or not b:
        return 0, "", ""
    if a == b:
        return 100, "high", "exact"
    ta, tb = a.split(), b.split()
    sa, sb = " ".join(sorted(ta)), " ".join(sorted(tb))
    if sa == sb:
        return 99, "high", "reversed"
    if a.replace(" ", "") == b.replace(" ", ""):
        return 98, "high", "nospace"
    sa_set, sb_set = set(ta), set(tb)
    shorter = sa_set if len(sa_set) <= len(sb_set) else sb_set
    longer = sb_set if len(sa_set) <= len(sb_set) else sa_set
    if shorter.issubset(longer) and len(shorter) >= 2:
        return 95, "medium", "token_subset"
    ratio = fuzz.ratio(sa, sb)
    if ratio >= FUZZY_THRESHOLD:
        return ratio, "low", "fuzzy"
    st = ta if len(ta) <= len(tb) else tb
    lt = tb if len(ta) <= len(tb) else ta
    matched = sum(1 for s in st if any(ngram_sim(s, l) >= NGRAM_THRESHOLD for l in lt))
    total = len(st)
    if total > 0 and matched >= max(2, total * 0.5):
        return int(80 + (matched / total) * 9), "low", "ngram"
    return 0, "", ""


# ── Excel 读取 ────────────────────────────────────────

def read_blacklist_excel(file_bytes: bytes) -> list[str]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["黑名单"] if "黑名单" in wb.sheetnames else wb.active
    names = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        v = row[0]
        if v and str(v).strip():
            names.append(str(v).strip())
    wb.close()
    return names


def scan_sheet(ws, max_col: int) -> tuple[list[list[Optional[str]]], list[dict]]:
    """扫描单个工作表，返回 grid + name_cells"""
    grid = []
    for r in range(1, ws.max_row + 1):
        row_data = []
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            row_data.append(str(v).strip() if v is not None else None)
        grid.append(row_data)

    name_cells = []
    for col_idx in NAME_COLUMNS:
        if col_idx > max_col:
            continue
        section = SECTIONS.get(col_idx, f"Col{col_idx}")
        for row_idx in range(1, ws.max_row):
            val = grid[row_idx][col_idx - 1] if col_idx - 1 < len(grid[row_idx]) else None
            if val and len(val) > 1:
                name_cells.append({
                    "row": row_idx + 1, "col": col_idx,
                    "name": val, "section": section,
                })
    return grid, name_cells


def check_sheet(name_cells: list[dict], bl: list[str]) -> dict:
    """对单个工作表执行匹配"""
    matches = []
    matched_cells = {}
    for cell in name_cells:
        best = (0, "", "")
        for bn in bl:
            result = match_score(cell["name"], bn)
            if result[0] > best[0]:
                best = result
        if best[0] > 0:
            key = f"{cell['row']},{cell['col']}"
            entry = {
                "row": cell["row"], "col": cell["col"],
                "cell": openpyxl.utils.get_column_letter(cell["col"]) + str(cell["row"]),
                "schedule_name": cell["name"],
                "blacklist_name": bl[[match_score(cell["name"], bn) for bn in bl].index(
                    max((match_score(cell["name"], bn) for bn in bl), key=lambda x: x[0])
                )] if best[0] > 0 else "",
                "score": best[0], "confidence": best[1], "method": best[2],
                "section": cell["section"],
            }
            # 修正 blacklist_name
            for bn in bl:
                s, c, m = match_score(cell["name"], bn)
                if s == best[0]:
                    entry["blacklist_name"] = bn
                    break
            matches.append(entry)
            matched_cells[key] = entry

    high = sum(1 for m in matches if m["confidence"] == "high")
    medium = sum(1 for m in matches if m["confidence"] == "medium")
    fuzzy = sum(1 for m in matches if m["method"] == "fuzzy")
    ngram = sum(1 for m in matches if m["method"] == "ngram")
    return {
        "total_scanned": len(name_cells),
        "total_matches": len(matches),
        "stats": {"high": high, "medium": medium, "fuzzy": fuzzy, "ngram": ngram},
        "matches": matches,
        "matched_cells": matched_cells,
    }


def check_workbook(file_bytes: bytes, bl: list[str]) -> dict:
    """检查整个工作簿的所有分表"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    max_col = min(max(ws.max_column for ws in wb.worksheets), 30)
    sheets_result = []
    total_scanned = 0
    total_matches = 0
    all_stats = {"high": 0, "medium": 0, "fuzzy": 0, "ngram": 0}

    for ws in wb.worksheets:
        grid, name_cells = scan_sheet(ws, max_col)
        result = check_sheet(name_cells, bl)
        result["sheet_name"] = ws.title
        result["grid"] = grid[:100]
        sheets_result.append(result)
        total_scanned += result["total_scanned"]
        total_matches += result["total_matches"]
        for k in all_stats:
            all_stats[k] += result["stats"][k]

    wb.close()
    return {
        "total_scanned": total_scanned,
        "total_matches": total_matches,
        "stats": all_stats,
        "sheets": sheets_result,
        "max_col": max_col,
    }


# ── Excel 导出（保留所有分表）────────────────────────

def generate_highlighted_workbook(file_bytes: bytes, sheets_result: list[dict]) -> bytes:
    """在原工作簿基础上标红匹配单元格，保留所有分表"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    red_font = openpyxl.styles.Font(color="FF0000", bold=True)
    red_fill = openpyxl.styles.PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    summary_inserted = False

    for sr in sheets_result:
        sheet_name = sr["sheet_name"]
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        mc = sr.get("matched_cells", {})

        for key, info in mc.items():
            r, c = info["row"], info["col"]
            cell = ws.cell(row=r, column=c)
            cell.font = red_font
            cell.fill = red_fill

        # 只在第一个有匹配的分表插入汇总行
        if not summary_inserted and sr["total_matches"] > 0:
            ws.insert_rows(1)
            s = sr["stats"]
            summary = (
                f"黑名单检查: 共{sr['total_matches']}个匹配 | "
                f"精确:{s['high']} 包含:{s['medium']} 模糊:{s['fuzzy']} N-gram:{s['ngram']}"
            )
            ws.cell(row=1, column=1, value=summary).font = openpyxl.styles.Font(
                color="FF0000", bold=True, size=12
            )
            summary_inserted = True

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb.close()
    return buf.getvalue()


# ── FastAPI ───────────────────────────────────────────

app = FastAPI(title="黑名单检查器")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
STATIC_DIR = Path(__file__).parent / "static"


@app.get("/")
async def root():
    return FileResponse(STATIC_DIR / "index.html")


# ── 黑名单管理 ────────────────────────────────────────

@app.get("/api/blacklist")
async def api_get_blacklist():
    return {"names": blacklist_names, "count": len(blacklist_names)}


@app.post("/api/blacklist/add")
async def api_add_blacklist(name: str = Form(...)):
    name = name.strip()
    if not name:
        return {"ok": False, "msg": "名字不能为空"}
    existing = _blacklist_set()
    added = []
    skipped = 0
    for n in [x.strip() for x in re.split(r"[,\n]", name) if x.strip()]:
        if _norm(n) not in existing:
            blacklist_names.append(n)
            existing.add(_norm(n))
            added.append(n)
        else:
            skipped += 1
    return {"ok": True, "added": added, "skipped": skipped, "total": len(blacklist_names)}


@app.post("/api/blacklist/upload")
async def api_upload_blacklist(file: UploadFile = File(...)):
    names = read_blacklist_excel(await file.read())
    existing = _blacklist_set()
    added = []
    for n in names:
        if _norm(n) not in existing:
            blacklist_names.append(n)
            existing.add(_norm(n))
            added.append(n)
    return {"ok": True, "added": added, "skipped": len(names) - len(added), "total": len(blacklist_names)}


@app.post("/api/blacklist/remove")
async def api_remove_blacklist(name: str = Form(...)):
    target = _norm(name)
    global blacklist_names
    before = len(blacklist_names)
    blacklist_names = [n for n in blacklist_names if _norm(n) != target]
    return {"ok": True, "removed": before - len(blacklist_names), "total": len(blacklist_names)}


@app.post("/api/blacklist/clear")
async def api_clear_blacklist():
    blacklist_names.clear()
    return {"ok": True, "total": 0}


# ── 批量检查 ──────────────────────────────────────────

@app.post("/api/batch-check")
async def api_batch_check(files: list[UploadFile] = File(...)):
    if not blacklist_names:
        return {"error": "黑名单为空，请先添加黑名单人员。"}
    bl = list(blacklist_names)
    results = []
    for f in files:
        data = await f.read()
        result = check_workbook(data, bl)
        result["filename"] = f.filename
        results.append(result)
    return {"results": results, "blacklist_count": len(bl)}


# ── 批量导出（zip）────────────────────────────────────

@app.post("/api/batch-export")
async def api_batch_export(files: list[UploadFile] = File(...)):
    if not blacklist_names:
        return Response(content='{"error":"黑名单为空"}', status_code=400)
    bl = list(blacklist_names)
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in files:
            data = await f.read()
            wb_result = check_workbook(data, bl)
            excel_bytes = generate_highlighted_workbook(data, wb_result["sheets"])
            out_name = Path(f.filename).stem + "_黑名单检查.xlsx"
            zf.writestr(out_name, excel_bytes)
    zip_buf.seek(0)
    encoded = quote("黑名单检查结果.zip")
    return Response(
        content=zip_buf.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8002)
