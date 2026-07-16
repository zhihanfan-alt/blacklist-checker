"""
Blacklist Checker / 分拣员工黑名单检查器
6-layer name matching + detailed reports + configurable columns.
"""

__version__ = "1.0.0"

import argparse
import io
import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from rapidfuzz import fuzz
from unidecode import unidecode

# ── Default configuration ──────────────────────────────

DEFAULT_CONFIG = {
    "name_columns": [1, 4, 8, 11, 15, 19, 23, 28],
    "sections": {
        "1": "Escaneadoras", "4": "Mesas", "8": "Name List",
        "11": "Name/Time1", "15": "Name/Time2", "19": "Morning List",
        "23": "Name3", "28": "Name4",
    },
    "fuzzy_threshold": 85,
    "ngram_threshold": 0.70,
    "blacklist_sheet": "黑名单",
}

# Runtime config — updated by load_config()
_cfg = dict(DEFAULT_CONFIG)


def load_config(path: str | None = None) -> dict:
    """Load config from JSON file, merged with defaults. Updates _cfg in place."""
    global _cfg
    _cfg = dict(DEFAULT_CONFIG)
    if path:
        with open(path, encoding="utf-8") as f:
            user = json.load(f)
        _cfg.update(user)
    return _cfg


# ── Matching engine ────────────────────────────────────

def _norm(name: str) -> str:
    name = name.strip().lower()
    name = unidecode(name)
    name = re.sub(r"[^\w\s]", "", name)
    return re.sub(r"\s+", " ", name).strip()


def ngram_sim(a: str, b: str) -> float:
    def ng(s, n): return {s[i:i+n] for i in range(len(s) - n + 1)}
    ba, bb = ng(a, 2), ng(b, 2)
    ta, tb = ng(a, 3), ng(b, 3)
    bs = len(ba & bb) / len(ba | bb) if (ba | bb) else 0
    ts = len(ta & tb) / len(ta | tb) if (ta | tb) else 0
    return (bs + ts) / 2


def match_score(sched: str, bl: str):
    if not sched or not bl:
        return 0, "", ""
    a, b = _norm(sched), _norm(bl)
    if not a or not b:
        return 0, "", ""
    if a == b:
        return 100, "high", "exact"
    ta, tb = a.split(), b.split()
    sa, sb = " ".join(sorted(ta)), " ".join(sorted(tb))
    if sa == sb:
        return 99, "high", "reversed"
    if a.replace(" ", "") == b.replace(" ", ""):
        return 98, "high", "nospace"
    sa_set, sb_set = set(ta), set(tb)
    shorter = sa_set if len(sa_set) <= len(sb_set) else sb_set
    longer = sb_set if len(sa_set) <= len(sb_set) else sa_set
    if shorter.issubset(longer) and len(shorter) >= 2:
        return 95, "medium", "token_subset"
    ratio = fuzz.ratio(sa, sb)
    if ratio >= _cfg["fuzzy_threshold"]:
        return ratio, "low", "fuzzy"
    st = ta if len(ta) <= len(tb) else tb
    lt = tb if len(ta) <= len(tb) else ta
    matched = sum(
        1 for s in st
        if any(ngram_sim(s, l) >= _cfg["ngram_threshold"] for l in lt)
    )
    total = len(st)
    if total > 0 and matched >= max(2, total * 0.5):
        return int(80 + (matched / total) * 9), "low", "ngram"
    return 0, "", ""


# ── Excel operations ───────────────────────────────────

def read_blacklist(path: str, sheet_name: str | None = None) -> list:
    wb = openpyxl.load_workbook(path, data_only=True)
    target = sheet_name or _cfg.get("blacklist_sheet", "黑名单")
    ws = wb[target] if target in wb.sheetnames else wb.active
    names = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        v = row[0]
        if v and str(v).strip():
            names.append(str(v).strip())
    wb.close()
    return names


def scan_sheet(ws, max_col: int):
    columns = _cfg["name_columns"]
    sections = {int(k): v for k, v in _cfg["sections"].items()}
    name_cells = []
    for col_idx in columns:
        if col_idx > max_col:
            continue
        section = sections.get(col_idx, f"Col{col_idx}")
        for row_idx in range(1, ws.max_row):
            val = ws.cell(row=row_idx + 1, column=col_idx).value
            if val and len(str(val).strip()) > 1:
                name_cells.append({
                    "row": row_idx + 1, "col": col_idx,
                    "name": str(val).strip(), "section": section,
                })
    return name_cells


def check_workbook(path: str, bl: list) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    max_col = min(max(ws.max_column for ws in wb.worksheets), 30)
    sheets = []
    total_scanned = 0
    total_matches = 0

    for ws in wb.worksheets:
        name_cells = scan_sheet(ws, max_col)
        matches = []
        for cell in name_cells:
            best_score, best_conf, best_method = 0, "", ""
            best_bl_name = ""
            for bn in bl:
                s, c, m = match_score(cell["name"], bn)
                if s > best_score:
                    best_score, best_conf, best_method, best_bl_name = s, c, m, bn
            if best_score > 0:
                matches.append({
                    "row": cell["row"], "col": cell["col"],
                    "cell": openpyxl.utils.get_column_letter(cell["col"]) + str(cell["row"]),
                    "schedule_name": cell["name"],
                    "blacklist_name": best_bl_name,
                    "score": best_score, "confidence": best_conf, "method": best_method,
                    "section": cell["section"],
                })

        sheets.append({
            "sheet_name": ws.title,
            "total_scanned": len(name_cells),
            "total_matches": len(matches),
            "matches": matches,
        })
        total_scanned += len(name_cells)
        total_matches += len(matches)

    wb.close()
    return {
        "filename": Path(path).name,
        "filepath": path,
        "total_scanned": total_scanned,
        "total_matches": total_matches,
        "sheets": sheets,
    }


def generate_highlighted(path: str, sheets_result: list) -> bytes:
    wb = openpyxl.load_workbook(path)
    red_font = Font(color="FF0000", bold=True)
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    summary_done = False

    for sr in sheets_result:
        if sr["sheet_name"] not in wb.sheetnames:
            continue
        ws = wb[sr["sheet_name"]]
        for m in sr["matches"]:
            cell = ws.cell(row=m["row"], column=m["col"])
            cell.font = red_font
            cell.fill = red_fill

        if not summary_done and sr["total_matches"] > 0:
            ws.insert_rows(1)
            ws.cell(row=1, column=1,
                    value=f"Blacklist matches: {sr['total_matches']} found / 黑名单检查: 共{sr['total_matches']}个匹配"
                    ).font = Font(color="FF0000", bold=True, size=12)
            summary_done = True

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


# ── CLI ────────────────────────────────────────────────

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="blacklist-checker",
        description="Check scheduling spreadsheets against an employee blacklist using 6-layer name matching.",
    )
    parser.add_argument("files", nargs="*", help="Excel files (first = blacklist, rest = schedules)")
    parser.add_argument("--version", action="version", version=f"%(prog)s {__version__}")
    parser.add_argument("--config", help="Path to JSON config file (see config.example.json)")
    parser.add_argument("--blacklist-sheet", help="Sheet name in the blacklist file (default: auto-detect '黑名单' or first sheet)")
    parser.add_argument("--output-dir", default="output", help="Output directory (default: output/)")
    parser.add_argument("--min-score", type=int, default=0, help="Minimum match score to include (0-100, default: 0)")
    return parser


def main():
    parser = build_parser()
    args = parser.parse_args()

    # Load config
    load_config(args.config)
    if args.blacklist_sheet:
        _cfg["blacklist_sheet"] = args.blacklist_sheet

    xlsx_files = args.files if args.files else sorted(
        str(f) for f in Path(".").glob("*.xlsx") if args.output_dir not in str(f)
    )

    if not xlsx_files:
        print("Error: No Excel files found. Please provide blacklist and schedule files.")
        print("用法: python checker.py blacklist.xlsx schedule1.xlsx schedule2.xlsx")
        sys.exit(1)

    # Identify blacklist file
    blacklist_path = None
    schedule_paths = []
    bl_sheet = args.blacklist_sheet

    for f in xlsx_files:
        fname = Path(f).name
        if "黑名单" in fname or "blacklist" in fname.lower():
            blacklist_path = f
        else:
            schedule_paths.append(f)

    # Fallback: check first file for matching sheet name
    if not blacklist_path and xlsx_files:
        try:
            wb = openpyxl.load_workbook(xlsx_files[0], read_only=True)
            target_sheet = bl_sheet or _cfg.get("blacklist_sheet", "黑名单")
            if target_sheet in wb.sheetnames:
                blacklist_path = xlsx_files[0]
                schedule_paths = xlsx_files[1:]
            wb.close()
        except Exception:
            pass

    if not blacklist_path:
        print("Error: Blacklist file not found (filename should contain '黑名单' or 'blacklist', "
              "or the first file should have a sheet named '黑名单')")
        sys.exit(1)
    if not schedule_paths:
        print("Error: No schedule files found")
        sys.exit(1)

    # Read blacklist
    print(f"[1/4] Reading blacklist: {Path(blacklist_path).name}")
    bl = read_blacklist(blacklist_path, bl_sheet)
    print(f"      Blacklist entries: {len(bl)}")
    for i, name in enumerate(bl[:20], 1):
        print(f"        {i:2d}. {name}")
    if len(bl) > 20:
        print(f"        ... {len(bl)} total")

    if not bl:
        print("Error: Blacklist is empty")
        sys.exit(1)

    # Check schedules
    print(f"\n[2/4] Checking schedules ({len(schedule_paths)} files)")
    all_results = []
    for sp in schedule_paths:
        print(f"      Checking: {Path(sp).name}")
        result = check_workbook(sp, bl)
        status = f"scanned {result['total_scanned']}, matched {result['total_matches']}"
        print(f"        -> {status}")
        for sheet in result["sheets"]:
            if sheet["total_matches"] > 0:
                print(f"          [{sheet['sheet_name']}] {sheet['total_matches']} matches")
                for m in sheet["matches"][:5]:
                    print(f"            {m['cell']}: {m['schedule_name']} vs {m['blacklist_name']} "
                          f"({m['score']}pts/{m['method']})")
                if sheet["total_matches"] > 5:
                    print(f"            ... +{sheet['total_matches'] - 5} more")
        all_results.append(result)

    # Filter by min-score
    if args.min_score > 0:
        for r in all_results:
            for s in r["sheets"]:
                s["matches"] = [m for m in s["matches"] if m["score"] >= args.min_score]
                s["total_matches"] = len(s["matches"])
            r["total_matches"] = sum(s["total_matches"] for s in r["sheets"])

    # Export matched files
    matched_files = [r for r in all_results if r["total_matches"] > 0]
    print(f"\n[3/4] Exporting results")
    output_dir = Path(args.output_dir)
    output_dir.mkdir(exist_ok=True)
    exported = []

    if matched_files:
        for r in matched_files:
            excel_bytes = generate_highlighted(r["filepath"], r["sheets"])
            out_name = Path(r["filename"]).stem + "_blacklist_check.xlsx"
            out_path = output_dir / out_name
            out_path.write_bytes(excel_bytes)
            exported.append(str(out_path))
            print(f"      -> {out_path} ({r['total_matches']} matches)")
    else:
        print("      No matches found in any file")

    # Summary
    print(f"\n[4/4] Summary")
    print(f"      Blacklist: {len(bl)} entries")
    print(f"      Files checked: {len(schedule_paths)}")
    print(f"      Files with matches: {len(matched_files)}")
    print(f"      Total matches: {sum(r['total_matches'] for r in all_results)}")
    print(f"      Exported: {len(exported)} files")

    summary = {
        "blacklist_count": len(bl),
        "blacklist_names": bl,
        "files_checked": len(schedule_paths),
        "files_matched": len(matched_files),
        "total_matches": sum(r["total_matches"] for r in all_results),
        "exported_files": exported,
        "results": [{
            "filename": r["filename"],
            "scanned": r["total_scanned"],
            "matches": r["total_matches"],
            "sheets": [{
                "name": s["sheet_name"],
                "scanned": s["total_scanned"],
                "matches": s["total_matches"],
                "details": s["matches"],
            } for s in r["sheets"]],
        } for r in all_results],
    }
    print(f"\n===JSON_REPORT===\n{json.dumps(summary, ensure_ascii=False)}")


if __name__ == "__main__":
    main()
